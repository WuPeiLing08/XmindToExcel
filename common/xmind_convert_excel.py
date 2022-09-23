from xmindparser import xmind_to_dict
import openpyxl
import os


def save_excel(records: list, tab_name: str, file_path: str):

    if os.path.exists(file_path):
        f = openpyxl.load_workbook(file_path)
    else:
        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"
        f = openpyxl.Workbook()
    sheet = f.create_sheet(tab_name)

    testcases_title = ["一级模块", "二级模块", "用例名称", "用例等级", "执行方式", "预置条件", "操作步骤", "期望结果"]
    for i in range(len(testcases_title)):
        sheet.cell(1, i+1, testcases_title[i])

    row = 1
    for record in records:
        row += 1
        try:
            sheet.cell(row, 1).value = record[0]
            sheet.cell(row, 2).value = record[1]
            sheet.cell(row, 3).value = record[-1]
            sheet.cell(row, 4).value = record[-2]
            sheet.cell(row, 5).value = "手动"
            sheet.cell(row, 6).value = ""
            steps = record[1:-3]
            steps_s = ""
            for i in range(len(steps)):
                steps_s = steps_s + str(i+1) + "." + steps[i] + "\n"
            sheet.cell(row, 7).value = steps_s.strip()
            sheet.cell(row, 8).value = record[-3]
        except Exception as e:
            print(e)
            continue
    f.save(file_path)


def get_x_title_list(records, record, tree):
    record.append(tree.get("title"))
    if "topics" in tree:
        for topic in tree.get("topics"):
            get_x_title_list(records, record.copy(), topic)
    else:
        records.append(record)


def x_convert_excel(xmind_file, convert_file):
    try:
        data_list = xmind_to_dict(xmind_file)
    except Exception as e:
        print(e)
    for data in data_list:
        tab_name = data.get("title")
        root = data.get("topic")
        records = []
        record = []
        get_x_title_list(records, record, root)
        save_excel(records, tab_name, convert_file)


if __name__ == "__main__":
    x_convert_excel("/Users/wpl/Downloads/xmind模版.xmind", "/Users/wpl/Downloads/excel00001模版.xlsx")



